"""
Forward Basis Matrix Generator (6M vs 3M)
Generate forward basis spread matrices showing 6M - 3M basis
across different forward start periods and swap tenors
Includes interactive 3D surface map visualization
Uses CUBIC SPLINE interpolation when available, falls back to linear
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
import sqlite3
from datetime import datetime, timedelta
import numpy as np
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from mpl_toolkits.mplot3d import Axes3D
from matplotlib import cm

# Try to import cubic spline - use linear interpolation as fallback
try:
    from scipy.interpolate import CubicSpline
    CUBIC_SPLINE_AVAILABLE = True
except ImportError:
    CUBIC_SPLINE_AVAILABLE = False
    print("⚠️ scipy not available - using linear interpolation")

class ForwardBasisMatrix:
    def __init__(self, root):
        self.root = root
        self.root.title("Forward Basis Matrix Generator (6M vs 3M)")
        self.root.geometry("1600x900")
        
        self.db_path = os.path.join(os.path.dirname(__file__), 'database', 'swap_rates.db')
        
        # AUD tenors (out to 30Y)
        self.aud_tenors = ['1Y', '2Y', '3Y', '4Y', '5Y', '7Y', '10Y', '12Y', '15Y', '20Y', '25Y', '30Y']
        
        # NZD tenors (out to 20Y)
        self.nzd_tenors = ['1Y', '2Y', '3Y', '4Y', '5Y', '7Y', '10Y', '12Y', '15Y', '20Y']
        
        # Forward start periods
        self.forward_periods = [
            ('1BD', 1/365),
            ('1m', 1/12),
            ('2m', 2/12),
            ('3m', 3/12),
            ('6m', 6/12),
            ('9m', 9/12),
            ('1y', 1),
            ('18m', 1.5),
            ('2y', 2),
            ('3y', 3),
            ('4y', 4),
            ('5y', 5),
            ('7y', 7),
            ('10y', 10),
            ('12y', 12),
            ('15y', 15),
            ('20y', 20)
        ]
        
        self.setup_ui()
    
    def setup_ui(self):
        # Header
        header = tk.Frame(self.root, bg='#2c3e50', padx=20, pady=15)
        header.pack(fill=tk.X)
        
        tk.Label(header, text="📊 Forward Basis Matrix Generator (6M vs 3M)", 
                bg='#2c3e50', fg='white', font=('Arial', 18, 'bold')).pack()
        
        tk.Label(header, text="Generate forward basis spread matrices showing 6M - 3M across forward periods", 
                bg='#2c3e50', fg='#ecf0f1', font=('Arial', 11)).pack()
        
        # Parameters frame
        params = tk.Frame(self.root, padx=20, pady=15)
        params.pack(fill=tk.X)
        
        # Currency
        currency_frame = tk.Frame(params)
        currency_frame.pack(side=tk.LEFT, padx=10)
        
        tk.Label(currency_frame, text="Currency:", font=('Arial', 11, 'bold')).pack(side=tk.LEFT, padx=5)
        self.currency_var = tk.StringVar(value='AUD')
        ttk.Combobox(currency_frame, textvariable=self.currency_var, 
                    values=['AUD', 'NZD'], width=8, state='readonly',
                    font=('Arial', 11)).pack(side=tk.LEFT, padx=5)
        
        # Date
        date_frame = tk.Frame(params)
        date_frame.pack(side=tk.LEFT, padx=10)
        
        tk.Label(date_frame, text="As of Date:", font=('Arial', 11, 'bold')).pack(side=tk.LEFT, padx=5)
        self.date_entry = DateEntry(date_frame, width=12, background='darkblue',
                                    foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        self.date_entry.set_date(datetime.now())
        self.date_entry.pack(side=tk.LEFT, padx=5)
        
        # Quick date buttons
        quick_dates_frame = tk.Frame(params)
        quick_dates_frame.pack(side=tk.LEFT, padx=10)
        
        tk.Button(quick_dates_frame, text="Today", command=lambda: self.set_quick_date(0),
                 bg='#95a5a6', fg='white', font=('Arial', 9),
                 padx=8, pady=4).pack(side=tk.LEFT, padx=2)
        
        tk.Button(quick_dates_frame, text="Yesterday", command=lambda: self.set_quick_date(1),
                 bg='#95a5a6', fg='white', font=('Arial', 9),
                 padx=8, pady=4).pack(side=tk.LEFT, padx=2)
        
        tk.Button(quick_dates_frame, text="Last Week", command=lambda: self.set_quick_date(7),
                 bg='#95a5a6', fg='white', font=('Arial', 9),
                 padx=8, pady=4).pack(side=tk.LEFT, padx=2)
        
        tk.Button(quick_dates_frame, text="Last Month", command=lambda: self.set_quick_date(30),
                 bg='#95a5a6', fg='white', font=('Arial', 9),
                 padx=8, pady=4).pack(side=tk.LEFT, padx=2)
        
        # Second row for buttons
        button_frame = tk.Frame(self.root, padx=20, pady=5)
        button_frame.pack(fill=tk.X)
        
        # Generate button
        tk.Button(button_frame, text="🔄 Generate Basis Matrix", command=self.generate_matrix,
                 bg='#27ae60', fg='white', font=('Arial', 12, 'bold'),
                 padx=20, pady=8, cursor='hand2').pack(side=tk.LEFT, padx=10)
        
        # Browse dates button
        tk.Button(button_frame, text="📅 Browse Dates", command=self.browse_available_dates,
                 bg='#9b59b6', fg='white', font=('Arial', 12, 'bold'),
                 padx=20, pady=8, cursor='hand2').pack(side=tk.LEFT, padx=5)
        
        # Export button
        tk.Button(button_frame, text="💾 Export to Excel", command=self.export_to_excel,
                 bg='#3498db', fg='white', font=('Arial', 12, 'bold'),
                 padx=20, pady=8, cursor='hand2').pack(side=tk.LEFT, padx=10)
        
        # 3D Surface button
        tk.Button(button_frame, text="🌐 Show 3D Surface", command=self.show_heat_map,
                 bg='#e74c3c', fg='white', font=('Arial', 12, 'bold'),
                 padx=20, pady=8, cursor='hand2').pack(side=tk.LEFT, padx=5)
        
        # Status
        self.status_var = tk.StringVar(value="Ready - Select currency and date, then Generate")
        tk.Label(self.root, textvariable=self.status_var, 
                font=('Arial', 10), fg='gray').pack(pady=5)
        
        # Matrix display frame with scrollbars
        matrix_container = tk.Frame(self.root)
        matrix_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Scrollbars
        h_scroll = tk.Scrollbar(matrix_container, orient=tk.HORIZONTAL)
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        
        v_scroll = tk.Scrollbar(matrix_container, orient=tk.VERTICAL)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Canvas for matrix
        self.canvas = tk.Canvas(matrix_container, 
                               xscrollcommand=h_scroll.set,
                               yscrollcommand=v_scroll.set,
                               bg='white')
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        h_scroll.config(command=self.canvas.xview)
        v_scroll.config(command=self.canvas.yview)
        
        # Frame inside canvas
        self.matrix_frame = tk.Frame(self.canvas, bg='white')
        self.canvas_window = self.canvas.create_window((0, 0), window=self.matrix_frame, anchor='nw')
        
        # Update scroll region when frame size changes
        self.matrix_frame.bind('<Configure>', 
                              lambda e: self.canvas.configure(scrollregion=self.canvas.bbox('all')))
        
        self.matrix_data = None
    
    def set_quick_date(self, days_ago):
        """Set date to X days ago"""
        target_date = datetime.now() - timedelta(days=days_ago)
        self.date_entry.set_date(target_date)
        self.status_var.set(f"Date set to {target_date.strftime('%Y-%m-%d')}")
    
    def get_available_dates(self, currency):
        """Get list of available dates in database for currency"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT DISTINCT date
                FROM swap_rates
                WHERE currency = ? 
                ORDER BY date DESC
                LIMIT 100
            """, (currency,))
            
            dates = [row[0] for row in cursor.fetchall()]
            conn.close()
            return dates
            
        except Exception as e:
            print(f"Error getting available dates: {e}")
            return []
    
    def browse_available_dates(self):
        """Show dialog to browse and select from available dates"""
        currency = self.currency_var.get()
        
        # Get available dates
        available_dates = self.get_available_dates(currency)
        
        if not available_dates:
            messagebox.showwarning("No Data", 
                f"No {currency} data found in database.\n\nPlease import data first.")
            return
        
        # Create dialog
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Available Dates - {currency}")
        dialog.geometry("400x600")
        
        # Header
        header = tk.Label(dialog, 
                         text=f"Available Dates for {currency}",
                         font=('Arial', 12, 'bold'), bg='#34495e', fg='white',
                         padx=10, pady=10)
        header.pack(fill=tk.X)
        
        info = tk.Label(dialog,
                       text=f"Found {len(available_dates)} dates with data\nDouble-click to select",
                       font=('Arial', 10), pady=5)
        info.pack()
        
        # Listbox with scrollbar
        list_frame = tk.Frame(dialog)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set,
                            font=('Courier', 10), height=25)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)
        
        # Populate listbox
        for date in available_dates:
            listbox.insert(tk.END, date)
        
        # Select current date if in list
        current_date = self.date_entry.get_date().strftime('%Y-%m-%d')
        if current_date in available_dates:
            idx = available_dates.index(current_date)
            listbox.select_set(idx)
            listbox.see(idx)
        
        def on_select(event=None):
            selection = listbox.curselection()
            if selection:
                selected_date = available_dates[selection[0]]
                self.date_entry.set_date(datetime.strptime(selected_date, '%Y-%m-%d'))
                dialog.destroy()
                self.status_var.set(f"Date set to {selected_date}")
        
        # Double-click to select
        listbox.bind('<Double-Button-1>', on_select)
        
        # Select button
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=10)
        
        tk.Button(btn_frame, text="Select Date", command=on_select,
                 bg='#27ae60', fg='white', font=('Arial', 11, 'bold'),
                 padx=20, pady=8).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Cancel", command=dialog.destroy,
                 bg='#95a5a6', fg='white', font=('Arial', 11, 'bold'),
                 padx=20, pady=8).pack(side=tk.LEFT, padx=5)
    
    def get_spot_curve(self, currency, date, floating_rate):
        """Get spot curve rates from database"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Determine fixing reference
            if currency == 'AUD':
                fixing = f'{floating_rate} BBSW'
            elif currency == 'NZD':
                fixing = f'{floating_rate} BKBM'
            else:
                fixing = floating_rate
            
            # Get rates for the date - handle both old and new formats
            cursor.execute("""
                SELECT tenor, rate
                FROM swap_rates
                WHERE date = ? AND currency = ? 
                AND (floating_rate = ? OR floating_rate LIKE ?)
                ORDER BY tenor
            """, (date, currency, fixing, f"{floating_rate}%"))
            
            rows = cursor.fetchall()
            conn.close()
            
            if not rows:
                return None
            
            # Convert to dictionary
            curve = {}
            for tenor, rate in rows:
                curve[tenor] = rate
            
            return curve
            
        except Exception as e:
            print(f"Error getting spot curve: {e}")
            return None
    
    def tenor_to_years(self, tenor):
        """Convert tenor string to years"""
        tenor = tenor.upper().strip()
        if tenor.endswith('Y'):
            return int(tenor[:-1])
        elif tenor.endswith('M'):
            return int(tenor[:-1]) / 12
        return 0
    
    def interpolate_rate(self, curve, years):
        """Cubic spline or linear interpolation of rates"""
        # Convert curve tenors to years
        points = [(self.tenor_to_years(t), r) for t, r in curve.items()]
        points.sort()
        
        if not points:
            return None
        
        # If exact match
        for y, r in points:
            if abs(y - years) < 0.01:
                return r
        
        # Need at least 2 points for interpolation
        if len(points) < 2:
            return points[0][1] if points else None
        
        # Extract x and y values
        x_points = np.array([p[0] for p in points])
        y_points = np.array([p[1] for p in points])
        
        # Check if target is within interpolation range
        if years < x_points[0]:
            # Flat extrapolation for points before first tenor
            return y_points[0]
        elif years > x_points[-1]:
            # Flat extrapolation for points after last tenor
            return y_points[-1]
        
        # Use cubic spline if available, otherwise linear
        if CUBIC_SPLINE_AVAILABLE and len(points) >= 3:
            try:
                # Create cubic spline with natural boundary conditions
                cs = CubicSpline(x_points, y_points, bc_type='natural')
                interpolated_rate = float(cs(years))
                return interpolated_rate
            except Exception as e:
                print(f"Cubic spline failed, using linear: {e}")
                # Fall through to linear interpolation
        
        # Linear interpolation fallback
        for i in range(len(points) - 1):
            y1, r1 = points[i]
            y2, r2 = points[i + 1]
            
            if y1 <= years <= y2:
                t = (years - y1) / (y2 - y1)
                return r1 + t * (r2 - r1)
        
        return None
    
    def calculate_forward_rate(self, spot_curve, forward_years, tenor_years):
        """Calculate forward starting swap rate"""
        # Get spot rates
        r1 = self.interpolate_rate(spot_curve, forward_years)
        r2 = self.interpolate_rate(spot_curve, forward_years + tenor_years)
        
        if r1 is None or r2 is None:
            return None
        
        # Forward rate calculation
        t1 = forward_years
        t2 = forward_years + tenor_years
        
        try:
            forward_rate = ((1 + r2)**t2 / (1 + r1)**t1)**(1/(t2 - t1)) - 1
            return forward_rate
        except:
            return None
    
    def calculate_forward_basis(self, curve_6m, curve_3m, forward_years, tenor_years):
        """Calculate forward basis spread (6M - 3M) in basis points"""
        # Calculate forward rates for both curves
        fwd_6m = self.calculate_forward_rate(curve_6m, forward_years, tenor_years)
        fwd_3m = self.calculate_forward_rate(curve_3m, forward_years, tenor_years)
        
        if fwd_6m is None or fwd_3m is None:
            return None
        
        # Basis = 6M - 3M (in basis points)
        basis = (fwd_6m - fwd_3m) * 10000
        return basis
    
    def generate_matrix(self):
        """Generate forward basis matrix"""
        currency = self.currency_var.get()
        date = self.date_entry.get_date().strftime('%Y-%m-%d')
        
        self.status_var.set(f"Generating {currency} 6M vs 3M forward basis matrix for {date}...")
        self.root.update()
        
        # Get both spot curves (6M and 3M)
        curve_6m = self.get_spot_curve(currency, date, '6M')
        curve_3m = self.get_spot_curve(currency, date, '3M')
        
        if not curve_6m or not curve_3m:
            # Try to find nearest available date
            available_dates = self.get_available_dates(currency)
            
            if available_dates:
                nearest = available_dates[0]
                msg = f"No complete {currency} 3M and 6M data found for {date}\n\n"
                msg += f"Most recent available date: {nearest}\n\n"
                msg += "Would you like to use the most recent date?"
                
                if messagebox.askyesno("Date Not Found", msg):
                    self.date_entry.set_date(datetime.strptime(nearest, '%Y-%m-%d'))
                    self.generate_matrix()
                    return
                else:
                    self.status_var.set(f"No data for {date}")
            else:
                messagebox.showerror("Error", 
                    f"No {currency} data found in database.\n\nPlease import both 3M and 6M data.")
                self.status_var.set("Error - No data")
            return
        
        # Determine tenors based on currency
        tenors = self.aud_tenors if currency == 'AUD' else self.nzd_tenors
        
        # Calculate forward basis spreads
        matrix = []
        for fwd_label, fwd_years in self.forward_periods:
            row = [fwd_label]
            for tenor in tenors:
                tenor_years = self.tenor_to_years(tenor)
                basis = self.calculate_forward_basis(curve_6m, curve_3m, fwd_years, tenor_years)
                row.append(basis)
            matrix.append(row)
        
        self.matrix_data = {
            'currency': currency,
            'date': date,
            'tenors': tenors,
            'matrix': matrix
        }
        
        # Display matrix
        self.display_matrix()
        
        self.status_var.set(f"Basis matrix generated for {currency} as of {date} - Showing 6M - 3M basis in bp")
    
    def display_matrix(self):
        """Display basis matrix in GUI with color coding"""
        # Clear previous matrix
        for widget in self.matrix_frame.winfo_children():
            widget.destroy()
        
        if not self.matrix_data:
            return
        
        currency = self.matrix_data['currency']
        date = self.matrix_data['date']
        tenors = self.matrix_data['tenors']
        matrix = self.matrix_data['matrix']
        
        # Title
        title = tk.Label(self.matrix_frame, 
                        text=f"{currency} Forward Basis Matrix (6M - 3M)",
                        font=('Arial', 14, 'bold'), bg='#34495e', fg='white',
                        padx=10, pady=8)
        title.grid(row=0, column=0, columnspan=len(tenors)+1, sticky='ew', padx=2, pady=2)
        
        # Date and subtitle
        subtitle = tk.Label(self.matrix_frame,
                           text=f"Date: {date} | Values in basis points (bp)",
                           font=('Arial', 11, 'bold'), bg='#ecf0f1',
                           padx=10, pady=5)
        subtitle.grid(row=1, column=0, columnspan=len(tenors)+1, sticky='ew', padx=2, pady=2)
        
        # Column headers
        tk.Label(self.matrix_frame, text="Fwd Period", 
                font=('Arial', 10, 'bold'), bg='#34495e', fg='white',
                width=10, pady=8).grid(row=2, column=0, sticky='ew', padx=1, pady=1)
        
        for col, tenor in enumerate(tenors, start=1):
            tk.Label(self.matrix_frame, text=tenor,
                    font=('Arial', 10, 'bold'), bg='#34495e', fg='white',
                    width=10, pady=8).grid(row=2, column=col, sticky='ew', padx=1, pady=1)
        
        # Data rows with color coding for basis spreads
        # Collect all basis values for color scaling
        all_basis = []
        for row in matrix:
            for val in row[1:]:
                if val is not None:
                    all_basis.append(val)
        
        if all_basis:
            min_basis = min(all_basis)
            max_basis = max(all_basis)
            mid_basis = (min_basis + max_basis) / 2
        else:
            min_basis = max_basis = mid_basis = 0
        
        for row_idx, row in enumerate(matrix, start=3):
            # Row header (forward period)
            tk.Label(self.matrix_frame, text=row[0],
                    font=('Arial', 10, 'bold'), bg='#ecf0f1',
                    width=10, pady=6).grid(row=row_idx, column=0, sticky='ew', padx=1, pady=1)
            
            # Data cells
            for col_idx, val in enumerate(row[1:], start=1):
                if val is None:
                    text = "-"
                    bg_color = '#f8f9fa'
                else:
                    text = f"{val:+.3f}"  # Show with + or - sign
                    
                    # Color coding for basis
                    # Positive basis (6M > 3M) - Normal, show in green shades
                    # Negative basis (6M < 3M) - Unusual, show in red shades
                    if val >= 0:
                        # Positive basis - green (normal situation)
                        if val > max_basis * 0.7:
                            bg_color = '#c8e6c9'  # Light green
                        elif val > max_basis * 0.4:
                            bg_color = '#e8f5e9'  # Very light green
                        else:
                            bg_color = '#f1f8e9'  # Pale green
                    else:
                        # Negative basis - red (unusual)
                        if val < min_basis * 0.7:
                            bg_color = '#ffcdd2'  # Light red
                        elif val < min_basis * 0.4:
                            bg_color = '#ffebee'  # Very light red
                        else:
                            bg_color = '#fff3e0'  # Pale orange
                
                tk.Label(self.matrix_frame, text=text,
                        font=('Courier', 10), bg=bg_color,
                        width=12, pady=6, relief='solid', borderwidth=1).grid(
                            row=row_idx, column=col_idx, sticky='ew', padx=1, pady=1)
    
    def show_heat_map(self):
        """Display 3D surface map of forward basis matrix"""
        if not self.matrix_data:
            messagebox.showwarning("Warning", "Please generate a matrix first")
            return
        
        currency = self.matrix_data['currency']
        date = self.matrix_data['date']
        tenors = self.matrix_data['tenors']
        matrix = self.matrix_data['matrix']
        
        # Create new window for 3D surface
        surface_window = tk.Toplevel(self.root)
        surface_window.title(f"3D Forward Basis Surface (6M - 3M) - {currency}")
        surface_window.geometry("1600x900")
        
        # Header
        header = tk.Frame(surface_window, bg='#2c3e50', padx=20, pady=10)
        header.pack(fill=tk.X)
        
        tk.Label(header, 
                text=f"3D Forward Basis Surface (6M - 3M) - {currency}",
                bg='#2c3e50', fg='white', font=('Arial', 16, 'bold')).pack()
        
        tk.Label(header,
                text=f"Date: {date} | Interactive 3D Surface - Click and drag to rotate",
                bg='#2c3e50', fg='#ecf0f1', font=('Arial', 11)).pack()
        
        # Prepare data for 3D surface
        forward_labels = [row[0] for row in matrix]
        
        # Extract numeric data (already in basis points)
        data_matrix = []
        for row in matrix:
            data_row = []
            for val in row[1:]:
                if val is None:
                    data_row.append(np.nan)
                else:
                    data_row.append(val)  # Already in basis points
            data_matrix.append(data_row)
        
        Z = np.array(data_matrix)
        
        # Create meshgrid for X (tenors) and Y (forward periods)
        X = np.arange(len(tenors))
        Y = np.arange(len(forward_labels))
        X, Y = np.meshgrid(X, Y)
        
        # Create 3D figure
        fig = plt.figure(figsize=(16, 10))
        ax = fig.add_subplot(111, projection='3d')
        
        # Create surface plot with custom colormap for basis
        # Positive basis (normal) = green, negative basis (unusual) = red
        surf = ax.plot_surface(X, Y, Z, cmap=cm.RdYlGn, 
                              linewidth=0.2, antialiased=True,
                              edgecolor='gray', alpha=0.9)
        
        # Customize axes
        ax.set_xlabel('Swap Tenor', fontsize=12, fontweight='bold', labelpad=10)
        ax.set_ylabel('Forward Start Period', fontsize=12, fontweight='bold', labelpad=10)
        ax.set_zlabel('Basis Spread (6M - 3M) in bp', fontsize=12, fontweight='bold', labelpad=10)
        
        # Set tick labels
        ax.set_xticks(np.arange(len(tenors)))
        ax.set_xticklabels(tenors, fontsize=9)
        ax.set_yticks(np.arange(len(forward_labels)))
        ax.set_yticklabels(forward_labels, fontsize=9)
        
        # Title
        ax.set_title(f'{currency} Forward Basis Surface (6M - 3M)\nDate: {date}',
                    fontsize=14, fontweight='bold', pad=20)
        
        # Add colorbar
        cbar = fig.colorbar(surf, ax=ax, shrink=0.5, aspect=10)
        cbar.set_label('Basis Spread (bp)', rotation=270, labelpad=20, fontsize=11)
        
        # Set viewing angle for better perspective
        ax.view_init(elev=25, azim=45)
        
        # Grid
        ax.grid(True, alpha=0.3)
        
        # Add zero plane if we have negative values
        if np.nanmin(Z) < 0:
            xx, yy = np.meshgrid(np.arange(len(tenors)), np.arange(len(forward_labels)))
            ax.plot_surface(xx, yy, np.zeros_like(Z), alpha=0.1, color='gray')
        
        plt.tight_layout()
        
        # Embed in tkinter
        canvas = FigureCanvasTkAgg(fig, master=surface_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Info frame
        info_frame = tk.Frame(surface_window, bg='#ecf0f1', pady=5)
        info_frame.pack(fill=tk.X)
        
        tk.Label(info_frame, 
                text="💡 Tip: Click and drag on the 3D surface to rotate. Use scroll wheel to zoom. Green = Positive basis (normal), Red = Negative basis (unusual)",
                bg='#ecf0f1', fg='#34495e', font=('Arial', 10, 'italic')).pack()
        
        # Buttons frame
        btn_frame = tk.Frame(surface_window, pady=10)
        btn_frame.pack(fill=tk.X)
        
        def save_surface():
            filename = filedialog.asksaveasfilename(
                defaultextension='.png',
                filetypes=[('PNG files', '*.png'), ('PDF files', '*.pdf'), ('All files', '*.*')],
                initialfile=f"Surface3D_FwdBasis_6v3_{currency}_{date}.png"
            )
            if filename:
                fig.savefig(filename, dpi=300, bbox_inches='tight')
                messagebox.showinfo("Success", f"3D surface saved to:\n{filename}")
        
        def reset_view():
            ax.view_init(elev=25, azim=45)
            canvas.draw()
        
        def top_view():
            ax.view_init(elev=90, azim=0)
            canvas.draw()
        
        def side_view():
            ax.view_init(elev=0, azim=0)
            canvas.draw()
        
        tk.Button(btn_frame, text="💾 Save 3D Surface", command=save_surface,
                 bg='#3498db', fg='white', font=('Arial', 11, 'bold'),
                 padx=20, pady=8).pack(side=tk.LEFT, padx=10)
        
        tk.Button(btn_frame, text="🔄 Reset View", command=reset_view,
                 bg='#f39c12', fg='white', font=('Arial', 11, 'bold'),
                 padx=20, pady=8).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="⬆️ Top View", command=top_view,
                 bg='#9b59b6', fg='white', font=('Arial', 11, 'bold'),
                 padx=20, pady=8).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="➡️ Side View", command=side_view,
                 bg='#16a085', fg='white', font=('Arial', 11, 'bold'),
                 padx=20, pady=8).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Close", command=surface_window.destroy,
                 bg='#95a5a6', fg='white', font=('Arial', 11, 'bold'),
                 padx=20, pady=8).pack(side=tk.RIGHT, padx=10)
    
    def export_to_excel(self):
        """Export basis matrix to Excel with formatting"""
        if not self.matrix_data:
            messagebox.showwarning("Warning", "Please generate a matrix first")
            return
        
        # Ask for file location
        filename = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')],
            initialfile=f"FwdBasisMatrix_6v3_{self.matrix_data['currency']}_{self.matrix_data['date']}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{self.matrix_data['currency']} 6v3 Basis"
            
            currency = self.matrix_data['currency']
            date = self.matrix_data['date']
            tenors = self.matrix_data['tenors']
            matrix = self.matrix_data['matrix']
            
            # Title
            ws.merge_cells('A1:' + chr(65 + len(tenors)) + '1')
            title_cell = ws['A1']
            title_cell.value = f"{currency} Forward Basis Matrix (6M - 3M)"
            title_cell.font = Font(bold=True, size=14, color='FFFFFF')
            title_cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Date
            ws.merge_cells('A2:' + chr(65 + len(tenors)) + '2')
            date_cell = ws['A2']
            date_cell.value = f"Date: {date} | Values in basis points (bp)"
            date_cell.font = Font(bold=True, size=11)
            date_cell.fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
            date_cell.alignment = Alignment(horizontal='center')
            
            # Headers
            header_row = 3
            ws.cell(header_row, 1, "Fwd Period").font = Font(bold=True, color='FFFFFF')
            ws.cell(header_row, 1).fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
            ws.cell(header_row, 1).alignment = Alignment(horizontal='center')
            
            for col, tenor in enumerate(tenors, start=2):
                cell = ws.cell(header_row, col, tenor)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Data with color coding
            all_basis = []
            for row in matrix:
                for val in row[1:]:
                    if val is not None:
                        all_basis.append(val)
            
            if all_basis:
                min_basis = min(all_basis)
                max_basis = max(all_basis)
            
            for row_idx, row in enumerate(matrix, start=4):
                # Row header
                cell = ws.cell(row_idx, 1, row[0])
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
                
                # Data
                for col_idx, val in enumerate(row[1:], start=2):
                    if val is None:
                        cell = ws.cell(row_idx, col_idx, "-")
                        cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                    else:
                        cell = ws.cell(row_idx, col_idx, val)
                        cell.number_format = '+0.000;-0.000'
                        
                        # Color coding for basis
                        if val >= 0:
                            if val > max_basis * 0.7:
                                color = 'C8E6C9'  # Light green
                            elif val > max_basis * 0.4:
                                color = 'E8F5E9'  # Very light green
                            else:
                                color = 'F1F8E9'  # Pale green
                        else:
                            if val < min_basis * 0.7:
                                color = 'FFCDD2'  # Light red
                            elif val < min_basis * 0.4:
                                color = 'FFEBEE'  # Very light red
                            else:
                                color = 'FFF3E0'  # Pale orange
                        
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # Column widths
            ws.column_dimensions['A'].width = 12
            for col in range(2, len(tenors) + 2):
                ws.column_dimensions[chr(64 + col)].width = 12
            
            wb.save(filename)
            messagebox.showinfo("Success", f"Basis matrix exported to:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export:\n{e}")

def main():
    root = tk.Tk()
    app = ForwardBasisMatrix(root)
    root.mainloop()

if __name__ == '__main__':
    main()
